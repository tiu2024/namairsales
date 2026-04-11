import io
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import DecimalField, ExpressionWrapper, F, Sum
from django.shortcuts import get_object_or_404, redirect, render

from accounts.decorators import accountant_required, admin_required
from accounts.forms import SalesmanCreateForm, SalesmanEditForm
from accounts.models import CustomUser

from .forms import AgentForm, AgentPaymentForm, ExpenditureForm, FinancialAccountForm, SaleForm, SupplierForm, SupplierPaymentForm
from .models import USD, UZS, Agent, AgentPayment, Expenditure, FinancialAccount, Sale, Supplier, SupplierPayment
from .services import edit_expenditure, link_sale_to_account, record_agent_payment, record_expenditure, record_supplier_payment, reverse_expenditure, unlink_sale_from_account


def index(request):
    if not request.user.is_authenticated:
        return redirect("accounts:login")
    if request.user.role == "ACCOUNTANT":
        return redirect("core:accountant_sales")
    return redirect("core:salesman_sales")


@login_required
def salesman_sales(request):
    if request.user.role == "ACCOUNTANT":
        return redirect("core:accountant_sales")
    # Parse filter params
    date_from_str = request.GET.get("date_from", "").strip()
    date_to_str   = request.GET.get("date_to", "").strip()
    currency      = request.GET.get("currency", "").strip()
    agent_id      = request.GET.get("agent", "").strip()
    supplier_id   = request.GET.get("supplier", "").strip()

    def parse_date(s):
        try:
            return datetime.strptime(s, "%d.%m.%Y").date()
        except ValueError:
            return None

    date_from = parse_date(date_from_str)
    date_to   = parse_date(date_to_str)

    # Base queryset
    all_sales = (
        Sale.objects.filter(salesman=request.user)
        .select_related("supplier", "agent")
        .order_by("-date", "-created_at")
    )

    # Apply filters
    if date_from:
        all_sales = all_sales.filter(date__gte=date_from)
    if date_to:
        all_sales = all_sales.filter(date__lte=date_to)
    if currency in ("UZS", "USD"):
        all_sales = all_sales.filter(sold_currency=currency)
    if agent_id.isdigit():
        all_sales = all_sales.filter(agent_id=int(agent_id))
    if supplier_id.isdigit():
        all_sales = all_sales.filter(supplier_id=int(supplier_id))

    # Aggregates (scoped to filtered queryset)
    total_sold_uzs = all_sales.filter(sold_currency="UZS").aggregate(
        t=Sum("sold_price"))["t"] or 0
    total_sold_usd = all_sales.filter(sold_currency="USD").aggregate(
        t=Sum("sold_price"))["t"] or 0

    _profit_expr = ExpressionWrapper(
        (F("sold_price") - F("acquired_price")) * F("quantity"),
        output_field=DecimalField(),
    )
    profit_uzs = (
        all_sales.filter(acquired_currency="UZS", sold_currency="UZS")
        .annotate(lp=_profit_expr)
        .aggregate(t=Sum("lp"))["t"] or 0
    )
    profit_usd = (
        all_sales.filter(acquired_currency="USD", sold_currency="USD")
        .annotate(lp=_profit_expr)
        .aggregate(t=Sum("lp"))["t"] or 0
    )

    paginator = Paginator(all_sales, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    params = request.GET.copy()
    params.pop("page", None)
    filter_query = params.urlencode()

    return render(request, "core/salesman_sales.html", {
        "sales": page_obj,
        "page_obj": page_obj,
        "form": SaleForm(),
        "total_sold_uzs": total_sold_uzs,
        "total_sold_usd": total_sold_usd,
        "profit_uzs": profit_uzs,
        "profit_usd": profit_usd,
        "filter_date_from": date_from_str,
        "filter_date_to": date_to_str,
        "filter_currency": currency,
        "filter_agent_id": agent_id,
        "filter_supplier_id": supplier_id,
        "all_agents": Agent.objects.order_by("name"),
        "all_suppliers": Supplier.objects.order_by("name"),
        "filter_query": filter_query,
    })


@login_required
def sale_form_get(request):
    if not request.headers.get("HX-Request"):
        return redirect("core:salesman_sales")
    return render(request, "core/partials/sale_form.html", {"form": SaleForm()})


@login_required
def sale_add(request):
    if not request.headers.get("HX-Request"):
        return redirect("core:salesman_sales")
    form = SaleForm(request.POST)
    if form.is_valid():
        sale = form.save(commit=False)
        sale.salesman = request.user
        sale.save()
        sale = Sale.objects.select_related("supplier", "agent").get(pk=sale.pk)
        return render(request, "core/partials/sale_row.html", {"sale": sale})
    response = render(request, "core/partials/sale_form.html", {"form": form}, status=400)
    response["HX-Reswap"] = "none"
    return response


@accountant_required
def accountant_sales(request):
    # Parse filter params
    date_from_str = request.GET.get("date_from", "").strip()
    date_to_str   = request.GET.get("date_to", "").strip()
    currency      = request.GET.get("currency", "").strip()
    agent_id      = request.GET.get("agent", "").strip()
    supplier_id   = request.GET.get("supplier", "").strip()
    salesman_id   = request.GET.get("salesman", "").strip()

    def parse_date(s):
        try:
            return datetime.strptime(s, "%d.%m.%Y").date()
        except ValueError:
            return None

    date_from = parse_date(date_from_str)
    date_to   = parse_date(date_to_str)

    all_sales = (
        Sale.objects.all()
        .select_related("supplier", "agent", "salesman", "financial_account")
        .order_by("-date", "-created_at")
    )

    if date_from:
        all_sales = all_sales.filter(date__gte=date_from)
    if date_to:
        all_sales = all_sales.filter(date__lte=date_to)
    if currency in ("UZS", "USD"):
        all_sales = all_sales.filter(sold_currency=currency)
    if agent_id.isdigit():
        all_sales = all_sales.filter(agent_id=int(agent_id))
    if supplier_id.isdigit():
        all_sales = all_sales.filter(supplier_id=int(supplier_id))
    if salesman_id.isdigit():
        all_sales = all_sales.filter(salesman_id=int(salesman_id))

    total_sold_uzs = all_sales.filter(sold_currency="UZS").aggregate(
        t=Sum("sold_price"))["t"] or 0
    total_sold_usd = all_sales.filter(sold_currency="USD").aggregate(
        t=Sum("sold_price"))["t"] or 0

    _profit_expr = ExpressionWrapper(
        (F("sold_price") - F("acquired_price")) * F("quantity"),
        output_field=DecimalField(),
    )
    profit_uzs = (
        all_sales.filter(acquired_currency="UZS", sold_currency="UZS")
        .annotate(lp=_profit_expr)
        .aggregate(t=Sum("lp"))["t"] or 0
    )
    profit_usd = (
        all_sales.filter(acquired_currency="USD", sold_currency="USD")
        .annotate(lp=_profit_expr)
        .aggregate(t=Sum("lp"))["t"] or 0
    )

    paginator = Paginator(all_sales, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    params = request.GET.copy()
    params.pop("page", None)
    filter_query = params.urlencode()

    return render(request, "core/accountant_sales.html", {
        "sales": page_obj,
        "page_obj": page_obj,
        "total_sold_uzs": total_sold_uzs,
        "total_sold_usd": total_sold_usd,
        "profit_uzs": profit_uzs,
        "profit_usd": profit_usd,
        "filter_date_from": date_from_str,
        "filter_date_to": date_to_str,
        "filter_currency": currency,
        "filter_agent_id": agent_id,
        "filter_supplier_id": supplier_id,
        "filter_salesman_id": salesman_id,
        "all_agents": Agent.objects.order_by("name"),
        "all_suppliers": Supplier.objects.order_by("name"),
        "all_salesmen": CustomUser.objects.filter(role="SALESMAN", is_active=True).order_by("phone_number"),
        "all_accounts": FinancialAccount.objects.filter(is_active=True).order_by("name"),
        "filter_query": filter_query,
    })


@accountant_required
def sale_link_account(request, pk):
    if not request.headers.get("HX-Request"):
        return redirect("core:accountant_sales")
    sale = get_object_or_404(Sale.objects.select_related("supplier", "agent", "salesman", "financial_account"), pk=pk)
    fa_id = request.POST.get("financial_account", "").strip()
    if not fa_id.isdigit():
        if not sale.financial_account_id:
            from django.http import HttpResponse
            r = HttpResponse(status=200)
            r["HX-Reswap"] = "none"
            return r
        unlink_sale_from_account(sale=sale, user=request.user)
    else:
        financial_account = get_object_or_404(FinancialAccount, pk=int(fa_id))
        try:
            link_sale_to_account(sale=sale, financial_account=financial_account, user=request.user)
        except ValueError as e:
            from django.http import HttpResponse
            r = HttpResponse(status=400)
            r["HX-Trigger"] = f'{{"toast": "{e}"}}'
            return r
    sale = Sale.objects.select_related("supplier", "agent", "salesman", "financial_account").get(pk=sale.pk)
    return render(request, "core/partials/accountant_sale_row.html", {
        "sale": sale,
        "all_accounts": FinancialAccount.objects.filter(is_active=True).order_by("name"),
    })


@login_required
def sale_edit(request, pk):
    sale = get_object_or_404(Sale.objects.select_related("financial_account"), pk=pk)
    if request.user.role != "ACCOUNTANT" and sale.salesman != request.user:
        return redirect("core:salesman_sales")

    linked_account = sale.financial_account  # snapshot before any changes

    if request.method == "POST":
        form = SaleForm(request.POST, instance=sale)
        if form.is_valid():
            with transaction.atomic():
                if sale.financial_account_id:
                    unlink_sale_from_account(sale=sale, user=request.user)
                updated_sale = form.save()
                if linked_account and linked_account.currency == updated_sale.sold_currency:
                    link_sale_to_account(
                        sale=updated_sale,
                        financial_account=linked_account,
                        user=request.user,
                    )
            if request.user.role == "ACCOUNTANT":
                return redirect("core:accountant_sales")
            return redirect("core:salesman_sales")
    else:
        form = SaleForm(instance=sale)

    return render(request, "core/sale_edit.html", {"form": form, "sale": sale})


@login_required
def sale_delete(request, pk):
    if request.method != "POST":
        return redirect("core:salesman_sales")
    sale = get_object_or_404(Sale, pk=pk)
    if request.user.role != "ACCOUNTANT" and sale.salesman != request.user:
        return redirect("core:salesman_sales")
    with transaction.atomic():
        if sale.financial_account_id:
            unlink_sale_from_account(sale=sale, user=request.user)
        sale.delete()
    if request.user.role == "ACCOUNTANT":
        return redirect("core:accountant_sales")
    return redirect("core:salesman_sales")


@accountant_required
def supplier_list(request):
    if request.method == "POST":
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("core:supplier_list")
    else:
        form = SupplierForm()

    q = request.GET.get("q", "").strip()
    suppliers = Supplier.objects.order_by("name")
    if q:
        suppliers = suppliers.filter(name__icontains=q)

    return render(request, "core/supplier_list.html", {
        "suppliers": suppliers,
        "form": form,
        "open": request.method == "POST" and not form.is_valid(),
        "q": q,
    })


@accountant_required
def supplier_detail(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)

    if request.method == "POST":
        payment_form = SupplierPaymentForm(request.POST)
        if payment_form.is_valid():
            cd = payment_form.cleaned_data
            record_supplier_payment(
                supplier=supplier,
                amount=cd["amount"],
                currency=cd["currency"],
                financial_account=cd["financial_account"],
                date=cd["date"],
                note=cd["note"],
                user=request.user,
            )
            return redirect("core:supplier_detail", pk=pk)
        payment_open = True
    else:
        payment_form = SupplierPaymentForm()
        payment_open = False

    filter_type = request.GET.get("type", "").strip()

    def _parse(s):
        try:
            return datetime.strptime(s, "%d.%m.%Y").date()
        except (ValueError, TypeError):
            return None

    date_from_str = request.GET.get("date_from", "").strip()
    date_to_str   = request.GET.get("date_to", "").strip()
    date_from = _parse(date_from_str)
    date_to   = _parse(date_to_str)

    # Sales queryset — filtered by type and/or date
    sales_qs = (
        Sale.objects.filter(supplier=supplier)
        .select_related("salesman")
        .order_by("-date", "-created_at")
    )
    if filter_type in (Sale.TICKET, Sale.UMRA, Sale.TOUR):
        sales_qs = sales_qs.filter(product_type=filter_type)
    if date_from:
        sales_qs = sales_qs.filter(date__gte=date_from)
    if date_to:
        sales_qs = sales_qs.filter(date__lte=date_to)

    # Payments queryset — filtered by date (no product type)
    pay_qs = (
        supplier.supplierpayment_set
        .select_related("financial_account", "created_by")
        .order_by("-date", "-created_at")
    )
    if date_from:
        pay_qs = pay_qs.filter(date__gte=date_from)
    if date_to:
        pay_qs = pay_qs.filter(date__lte=date_to)

    # Summary card balance
    cost_expr = ExpressionWrapper(F("acquired_price") * F("quantity"), output_field=DecimalField())
    if filter_type in (Sale.TICKET, Sale.UMRA, Sale.TOUR):
        # Type filter: show acquisition costs only (payments have no type)
        balance_uzs = sales_qs.filter(acquired_currency=UZS).annotate(c=cost_expr).aggregate(t=Sum("c"))["t"] or 0
        balance_usd = sales_qs.filter(acquired_currency=USD).annotate(c=cost_expr).aggregate(t=Sum("c"))["t"] or 0
    elif date_from or date_to:
        # Date filter: acquisitions minus payments for the period, including initial balance
        balance_uzs = (
            (sales_qs.filter(acquired_currency=UZS).annotate(c=cost_expr).aggregate(t=Sum("c"))["t"] or 0)
            - (pay_qs.filter(currency=UZS).aggregate(t=Sum("amount"))["t"] or 0)
            - supplier.initial_balance_uzs
        )
        balance_usd = (
            (sales_qs.filter(acquired_currency=USD).annotate(c=cost_expr).aggregate(t=Sum("c"))["t"] or 0)
            - (pay_qs.filter(currency=USD).aggregate(t=Sum("amount"))["t"] or 0)
            - supplier.initial_balance_usd
        )
    else:
        balance_uzs = supplier.balance_uzs()
        balance_usd = supplier.balance_usd()

    # Combined rows for main table
    if filter_type in (Sale.TICKET, Sale.UMRA, Sale.TOUR):
        rows = [{"kind": "sale", "obj": s} for s in sales_qs]
    else:
        rows = (
            [{"kind": "sale", "obj": s} for s in sales_qs]
            + [{"kind": "payment", "obj": p} for p in pay_qs]
        )
        rows.sort(key=lambda x: (x["obj"].date, x["obj"].created_at), reverse=True)

    paginator = Paginator(rows, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "core/supplier_detail.html", {
        "supplier": supplier,
        "page_obj": page_obj,
        "filter_type": filter_type,
        "balance_uzs": balance_uzs,
        "balance_usd": balance_usd,
        "payment_form": payment_form,
        "payment_open": payment_open,
        "date_from_str": date_from_str,
        "date_to_str": date_to_str,
    })


@accountant_required
def agent_list(request):
    if request.method == "POST":
        form = AgentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("core:agent_list")
    else:
        form = AgentForm()

    q = request.GET.get("q", "").strip()
    agents = Agent.objects.order_by("name")
    if q:
        agents = agents.filter(name__icontains=q)

    return render(request, "core/agent_list.html", {
        "agents": agents,
        "form": form,
        "open": request.method == "POST" and not form.is_valid(),
        "q": q,
    })


@accountant_required
def agent_detail(request, pk):
    agent = get_object_or_404(Agent, pk=pk)

    if request.method == "POST":
        payment_form = AgentPaymentForm(request.POST)
        if payment_form.is_valid():
            cd = payment_form.cleaned_data
            record_agent_payment(
                agent=agent,
                amount=cd["amount"],
                currency=cd["currency"],
                financial_account=cd["financial_account"],
                date=cd["date"],
                note=cd["note"],
                user=request.user,
            )
            return redirect("core:agent_detail", pk=pk)
        payment_open = True
    else:
        payment_form = AgentPaymentForm()
        payment_open = False

    filter_type = request.GET.get("type", "").strip()

    def _parse(s):
        try:
            return datetime.strptime(s, "%d.%m.%Y").date()
        except (ValueError, TypeError):
            return None

    date_from_str = request.GET.get("date_from", "").strip()
    date_to_str   = request.GET.get("date_to", "").strip()
    date_from = _parse(date_from_str)
    date_to   = _parse(date_to_str)

    # Sales queryset — only AGENT-type sales for this agent
    sales_qs = (
        Sale.objects.filter(agent=agent)
        .select_related("salesman")
        .order_by("-date", "-created_at")
    )
    if filter_type in (Sale.TICKET, Sale.UMRA, Sale.TOUR):
        sales_qs = sales_qs.filter(product_type=filter_type)
    if date_from:
        sales_qs = sales_qs.filter(date__gte=date_from)
    if date_to:
        sales_qs = sales_qs.filter(date__lte=date_to)

    # Payments queryset
    pay_qs = (
        agent.agentpayment_set
        .select_related("financial_account", "created_by")
        .order_by("-date", "-created_at")
    )
    if date_from:
        pay_qs = pay_qs.filter(date__gte=date_from)
    if date_to:
        pay_qs = pay_qs.filter(date__lte=date_to)

    # Balance — agent formula: initial_balance + sales(sold_price) - payments
    sold_expr = ExpressionWrapper(F("sold_price") * F("quantity"), output_field=DecimalField())
    if filter_type in (Sale.TICKET, Sale.UMRA, Sale.TOUR):
        balance_uzs = sales_qs.filter(sold_currency=UZS).annotate(c=sold_expr).aggregate(t=Sum("c"))["t"] or 0
        balance_usd = sales_qs.filter(sold_currency=USD).annotate(c=sold_expr).aggregate(t=Sum("c"))["t"] or 0
    elif date_from or date_to:
        balance_uzs = (
            agent.initial_balance_uzs
            + (sales_qs.filter(sold_currency=UZS).annotate(c=sold_expr).aggregate(t=Sum("c"))["t"] or 0)
            - (pay_qs.filter(currency=UZS).aggregate(t=Sum("amount"))["t"] or 0)
        )
        balance_usd = (
            agent.initial_balance_usd
            + (sales_qs.filter(sold_currency=USD).annotate(c=sold_expr).aggregate(t=Sum("c"))["t"] or 0)
            - (pay_qs.filter(currency=USD).aggregate(t=Sum("amount"))["t"] or 0)
        )
    else:
        balance_uzs = agent.balance_uzs()
        balance_usd = agent.balance_usd()

    # Combined rows
    if filter_type in (Sale.TICKET, Sale.UMRA, Sale.TOUR):
        rows = [{"kind": "sale", "obj": s} for s in sales_qs]
    else:
        rows = (
            [{"kind": "sale", "obj": s} for s in sales_qs]
            + [{"kind": "payment", "obj": p} for p in pay_qs]
        )
        rows.sort(key=lambda x: (x["obj"].date, x["obj"].created_at), reverse=True)

    paginator = Paginator(rows, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "core/agent_detail.html", {
        "agent": agent,
        "page_obj": page_obj,
        "filter_type": filter_type,
        "balance_uzs": balance_uzs,
        "balance_usd": balance_usd,
        "payment_form": payment_form,
        "payment_open": payment_open,
        "date_from_str": date_from_str,
        "date_to_str": date_to_str,
    })


@accountant_required
def agent_edit(request, pk):
    agent = get_object_or_404(Agent, pk=pk)
    if request.method == "POST":
        form = AgentForm(request.POST, instance=agent)
        if form.is_valid():
            form.save()
            return redirect("core:agent_detail", pk=pk)
    else:
        form = AgentForm(instance=agent)
    return render(request, "core/agent_edit.html", {"form": form, "agent": agent})


@admin_required
def agent_delete(request, pk):
    agent = get_object_or_404(Agent, pk=pk)
    if request.method == "POST":
        agent.delete()
        return redirect("core:agent_list")
    return render(request, "core/agent_confirm_delete.html", {"object": agent})


@accountant_required
def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == "POST":
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect("core:supplier_detail", pk=pk)
    else:
        form = SupplierForm(instance=supplier)
    return render(request, "core/supplier_edit.html", {"form": form, "supplier": supplier})


@admin_required
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == "POST":
        supplier.delete()
        return redirect("core:supplier_list")
    return render(request, "core/supplier_confirm_delete.html", {"object": supplier})


@accountant_required
def expenditure_list(request):
    form = ExpenditureForm()
    form_open = False

    if request.method == "POST":
        form = ExpenditureForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            record_expenditure(
                amount=cd["amount"],
                currency=cd["currency"],
                financial_account=cd["financial_account"],
                date=cd["date"],
                description=cd["description"],
                user=request.user,
            )
            return redirect("core:expenditure_list")
        form_open = True

    expenditures = (
        Expenditure.objects.filter(is_active=True)
        .select_related("financial_account", "registered_by")
        .order_by("-date", "-created_at")
    )
    return render(request, "core/expenditure_list.html", {
        "expenditures": expenditures,
        "form": form,
        "form_open": form_open,
    })


@login_required
def expenditure_edit(request, pk):
    expenditure = get_object_or_404(Expenditure, pk=pk, is_active=True)
    if request.method == "POST":
        form = ExpenditureForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            edit_expenditure(
                expenditure=expenditure,
                amount=cd["amount"],
                currency=cd["currency"],
                financial_account=cd["financial_account"],
                date=cd["date"],
                description=cd["description"],
                user=request.user,
            )
            return redirect("core:expenditure_list")
    else:
        form = ExpenditureForm(initial={
            "amount": expenditure.amount,
            "currency": expenditure.currency,
            "financial_account": expenditure.financial_account,
            "date": expenditure.date.strftime("%d.%m.%Y"),
            "description": expenditure.description,
        })
    return render(request, "core/expenditure_edit.html", {"form": form, "expenditure": expenditure})


@login_required
def expenditure_archive(request, pk):
    expenditure = get_object_or_404(Expenditure, pk=pk, is_active=True)
    if request.method == "POST":
        reverse_expenditure(expenditure=expenditure, user=request.user)
        expenditure.is_active = False
        expenditure.save(update_fields=["is_active"])
        return redirect("core:expenditure_list")


@accountant_required
def financial_account_list(request):
    if request.method == "POST":
        form = FinancialAccountForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("core:financial_account_list")
    else:
        form = FinancialAccountForm()

    q = request.GET.get("q", "").strip()
    accounts = FinancialAccount.objects.filter(is_active=True).order_by("name")
    if q:
        accounts = accounts.filter(name__icontains=q)
    return render(request, "core/financial_account_list.html", {
        "accounts": accounts,
        "form": form,
        "open": request.method == "POST" and not form.is_valid(),
        "q": q,
    })


@accountant_required
def financial_account_detail(request, pk):
    account = get_object_or_404(FinancialAccount, pk=pk)

    date_from_str = request.GET.get("date_from", "").strip()
    date_to_str   = request.GET.get("date_to", "").strip()

    def parse_date(s):
        try:
            return datetime.strptime(s, "%d.%m.%Y").date()
        except ValueError:
            return None

    date_from = parse_date(date_from_str)
    date_to   = parse_date(date_to_str)

    sales = Sale.objects.filter(financial_account=account).select_related("salesman", "supplier", "agent")
    supplier_payments = SupplierPayment.objects.filter(financial_account=account).select_related("supplier", "created_by")
    agent_payments = AgentPayment.objects.filter(financial_account=account).select_related("agent", "created_by")
    expenditures = Expenditure.objects.filter(financial_account=account).select_related("registered_by")

    if date_from:
        sales = sales.filter(date__gte=date_from)
        supplier_payments = supplier_payments.filter(date__gte=date_from)
        agent_payments = agent_payments.filter(date__gte=date_from)
        expenditures = expenditures.filter(date__gte=date_from)
    if date_to:
        sales = sales.filter(date__lte=date_to)
        supplier_payments = supplier_payments.filter(date__lte=date_to)
        agent_payments = agent_payments.filter(date__lte=date_to)
        expenditures = expenditures.filter(date__lte=date_to)

    rows = []
    for obj in sales:
        rows.append({
            "date": obj.date,
            "created_at": obj.created_at,
            "kind": "sale",
            "label": "Sotuv",
            "who": obj.destination,
            "note": obj.commentary,
            "amount": obj.sold_price,
            "currency": obj.sold_currency,
            "sign": "+",
            "actor": obj.salesman.get_full_name() or obj.salesman.phone_number,
        })
    for obj in supplier_payments:
        rows.append({
            "date": obj.date,
            "created_at": obj.created_at,
            "kind": "supplier_payment",
            "label": "Yetkazib beruvchiga to'lov",
            "who": obj.supplier.name,
            "note": obj.note,
            "amount": obj.amount,
            "currency": obj.currency,
            "sign": "-",
            "actor": obj.created_by.get_full_name() or obj.created_by.phone_number,
        })
    for obj in agent_payments:
        rows.append({
            "date": obj.date,
            "created_at": obj.created_at,
            "kind": "agent_payment",
            "label": "Agentga to'lov",
            "who": obj.agent.name,
            "note": obj.note,
            "amount": obj.amount,
            "currency": obj.currency,
            "sign": "-",
            "actor": obj.created_by.get_full_name() or obj.created_by.phone_number,
        })
    for obj in expenditures:
        rows.append({
            "date": obj.date,
            "created_at": obj.created_at,
            "kind": "expenditure",
            "label": "Xarajat",
            "who": obj.description,
            "note": "",
            "amount": obj.amount,
            "currency": obj.currency,
            "sign": "-",
            "actor": obj.registered_by.get_full_name() or obj.registered_by.phone_number,
        })

    rows.sort(key=lambda x: (x["date"], x["created_at"]), reverse=True)

    paginator = Paginator(rows, 30)
    page_obj = paginator.get_page(request.GET.get("page"))

    params = request.GET.copy()
    params.pop("page", None)
    filter_query = params.urlencode()

    return render(request, "core/financial_account_detail.html", {
        "account": account,
        "page_obj": page_obj,
        "date_from_str": date_from_str,
        "date_to_str": date_to_str,
        "filter_query": filter_query,
    })


@accountant_required
def financial_account_edit(request, pk):
    account = get_object_or_404(FinancialAccount, pk=pk)
    if request.method == "POST":
        form = FinancialAccountForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            return redirect("core:financial_account_detail", pk=pk)
    else:
        form = FinancialAccountForm(instance=account)
    return render(request, "core/financial_account_edit.html", {"form": form, "account": account})


@admin_required
def financial_account_archive(request, pk):
    account = get_object_or_404(FinancialAccount, pk=pk)
    if request.method == "POST":
        account.is_active = False
        account.save(update_fields=["is_active"])
        return redirect("core:financial_account_list")


@admin_required
def salesman_list(request):
    if request.method == "POST":
        form = SalesmanCreateForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("core:salesman_list")
    else:
        form = SalesmanCreateForm()

    salesmen = CustomUser.objects.filter(role="SALESMAN", is_active=True).order_by("first_name", "last_name")
    return render(request, "core/salesman_list.html", {
        "salesmen": salesmen,
        "form": form,
        "open": request.method == "POST" and not form.is_valid(),
    })


@admin_required
def salesman_edit(request, pk):
    salesman = get_object_or_404(CustomUser, pk=pk, role="SALESMAN")
    if request.method == "POST":
        form = SalesmanEditForm(request.POST, instance=salesman)
        if form.is_valid():
            form.save()
            return redirect("core:salesman_list")
    else:
        form = SalesmanEditForm(instance=salesman)
    return render(request, "core/salesman_edit.html", {"form": form, "salesman": salesman})


@admin_required
def salesman_deactivate(request, pk):
    salesman = get_object_or_404(CustomUser, pk=pk, role="SALESMAN")
    if request.method == "POST":
        salesman.is_active = False
        salesman.save(update_fields=["is_active"])
        return redirect("core:salesman_list")


# ── Excel export helpers ────────────────────────────────────────────────────

def _excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _parse_date(s):
    try:
        return datetime.strptime(s, "%d.%m.%Y").date()
    except (ValueError, TypeError):
        return None


@accountant_required
def accountant_sales_export(request):
    date_from = _parse_date(request.GET.get("date_from", ""))
    date_to   = _parse_date(request.GET.get("date_to", ""))
    currency    = request.GET.get("currency", "").strip()
    agent_id    = request.GET.get("agent", "").strip()
    supplier_id = request.GET.get("supplier", "").strip()
    salesman_id = request.GET.get("salesman", "").strip()

    qs = Sale.objects.all().select_related("supplier", "agent", "salesman", "financial_account").order_by("-date", "-created_at")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if currency in ("UZS", "USD"):
        qs = qs.filter(sold_currency=currency)
    if agent_id.isdigit():
        qs = qs.filter(agent_id=int(agent_id))
    if supplier_id.isdigit():
        qs = qs.filter(supplier_id=int(supplier_id))
    if salesman_id.isdigit():
        qs = qs.filter(salesman_id=int(salesman_id))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sotuvlar"
    ws.append(["Sana", "Sotuvchi", "Yetkazib beruvchi", "Turi", "Yo'nalish",
                "Miqdor", "Tannarx", "Tan. valyuta", "Sotish narxi", "Sot. valyuta",
                "Foyda", "Foyda valyuta", "Mijoz", "Hisob"])
    for s in qs:
        customer = s.agent.name if s.agent else s.customer_name
        if s.acquired_currency == s.sold_currency:
            profit = float((s.sold_price - s.acquired_price) * s.quantity)
            profit_currency = s.sold_currency
        else:
            profit = ""
            profit_currency = ""
        ws.append([
            s.date.strftime("%d.%m.%Y"),
            s.salesman.get_full_name() or s.salesman.phone_number,
            s.supplier.name,
            s.get_product_type_display(),
            s.destination,
            s.quantity,
            float(s.acquired_price),
            s.acquired_currency,
            float(s.sold_price),
            s.sold_currency,
            profit,
            profit_currency,
            customer,
            s.financial_account.name if s.financial_account else "",
        ])
    return _excel_response(wb, "sotuvlar.xlsx")


@login_required
def salesman_sales_export(request):
    if request.user.role == "ACCOUNTANT":
        return redirect("core:accountant_sales_export")
    date_from   = _parse_date(request.GET.get("date_from", ""))
    date_to     = _parse_date(request.GET.get("date_to", ""))
    currency    = request.GET.get("currency", "").strip()
    agent_id    = request.GET.get("agent", "").strip()
    supplier_id = request.GET.get("supplier", "").strip()

    qs = Sale.objects.filter(salesman=request.user).select_related("supplier", "agent").order_by("-date", "-created_at")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if currency in ("UZS", "USD"):
        qs = qs.filter(sold_currency=currency)
    if agent_id.isdigit():
        qs = qs.filter(agent_id=int(agent_id))
    if supplier_id.isdigit():
        qs = qs.filter(supplier_id=int(supplier_id))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sotuvlarim"
    ws.append(["Sana", "Yetkazib beruvchi", "Turi", "Yo'nalish",
                "Miqdor", "Tannarx", "Tan. valyuta", "Sotish narxi", "Sot. valyuta",
                "Foyda", "Foyda valyuta", "Mijoz"])
    for s in qs:
        customer = s.agent.name if s.agent else s.customer_name
        if s.acquired_currency == s.sold_currency:
            profit = float((s.sold_price - s.acquired_price) * s.quantity)
            profit_currency = s.sold_currency
        else:
            profit = ""
            profit_currency = ""
        ws.append([
            s.date.strftime("%d.%m.%Y"),
            s.supplier.name,
            s.get_product_type_display(),
            s.destination,
            s.quantity,
            float(s.acquired_price),
            s.acquired_currency,
            float(s.sold_price),
            s.sold_currency,
            profit,
            profit_currency,
            customer,
        ])
    return _excel_response(wb, "sotuvlarim.xlsx")


@accountant_required
def agent_detail_export(request, pk):
    agent = get_object_or_404(Agent, pk=pk)
    filter_type = request.GET.get("type", "").strip()
    date_from   = _parse_date(request.GET.get("date_from", ""))
    date_to     = _parse_date(request.GET.get("date_to", ""))

    sales_qs = Sale.objects.filter(agent=agent).select_related("salesman").order_by("-date", "-created_at")
    if filter_type in (Sale.TICKET, Sale.UMRA, Sale.TOUR):
        sales_qs = sales_qs.filter(product_type=filter_type)
    if date_from:
        sales_qs = sales_qs.filter(date__gte=date_from)
    if date_to:
        sales_qs = sales_qs.filter(date__lte=date_to)

    pay_qs = agent.agentpayment_set.select_related("created_by").order_by("-date", "-created_at")
    if date_from:
        pay_qs = pay_qs.filter(date__gte=date_from)
    if date_to:
        pay_qs = pay_qs.filter(date__lte=date_to)

    if filter_type in (Sale.TICKET, Sale.UMRA, Sale.TOUR):
        rows = [{"kind": "sale", "obj": s} for s in sales_qs]
    else:
        rows = (
            [{"kind": "sale", "obj": s} for s in sales_qs]
            + [{"kind": "payment", "obj": p} for p in pay_qs]
        )
        rows.sort(key=lambda x: (x["obj"].date, x["obj"].created_at), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = agent.name[:31]
    ws.append(["Sana", "Turi", "Yo'nalish / Izoh", "Miqdor",
                "Sotuv (so'm)", "Sotuv ($)", "To'lov (so'm)", "To'lov ($)", "Kim"])
    for row in rows:
        obj = row["obj"]
        if row["kind"] == "sale":
            ws.append([
                obj.date.strftime("%d.%m.%Y"),
                obj.get_product_type_display(),
                obj.destination,
                obj.quantity,
                float(obj.sold_price) if obj.sold_currency == UZS else "",
                float(obj.sold_price) if obj.sold_currency == USD else "",
                "",
                "",
                obj.salesman.get_full_name() or obj.salesman.phone_number,
            ])
        else:
            ws.append([
                obj.date.strftime("%d.%m.%Y"),
                "To'lov",
                obj.note or "",
                "",
                "",
                "",
                float(obj.amount) if obj.currency == UZS else "",
                float(obj.amount) if obj.currency == USD else "",
                obj.created_by.get_full_name() or obj.created_by.phone_number,
            ])
    return _excel_response(wb, f"agent-{agent.pk}.xlsx")


@accountant_required
def supplier_detail_export(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    filter_type = request.GET.get("type", "").strip()
    date_from   = _parse_date(request.GET.get("date_from", ""))
    date_to     = _parse_date(request.GET.get("date_to", ""))

    sales_qs = Sale.objects.filter(supplier=supplier).select_related("salesman").order_by("-date", "-created_at")
    if filter_type in (Sale.TICKET, Sale.UMRA, Sale.TOUR):
        sales_qs = sales_qs.filter(product_type=filter_type)
    if date_from:
        sales_qs = sales_qs.filter(date__gte=date_from)
    if date_to:
        sales_qs = sales_qs.filter(date__lte=date_to)

    pay_qs = supplier.supplierpayment_set.select_related("created_by").order_by("-date", "-created_at")
    if date_from:
        pay_qs = pay_qs.filter(date__gte=date_from)
    if date_to:
        pay_qs = pay_qs.filter(date__lte=date_to)

    if filter_type in (Sale.TICKET, Sale.UMRA, Sale.TOUR):
        rows = [{"kind": "sale", "obj": s} for s in sales_qs]
    else:
        rows = (
            [{"kind": "sale", "obj": s} for s in sales_qs]
            + [{"kind": "payment", "obj": p} for p in pay_qs]
        )
        rows.sort(key=lambda x: (x["obj"].date, x["obj"].created_at), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = supplier.name[:31]
    ws.append(["Sana", "Turi", "Yo'nalish / Izoh", "Miqdor",
                "Tannarx (so'm)", "Tannarx ($)", "To'lov (so'm)", "To'lov ($)", "Kim"])
    for row in rows:
        obj = row["obj"]
        if row["kind"] == "sale":
            ws.append([
                obj.date.strftime("%d.%m.%Y"),
                obj.get_product_type_display(),
                obj.destination,
                obj.quantity,
                float(obj.total_cost) if obj.acquired_currency == UZS else "",
                float(obj.total_cost) if obj.acquired_currency == USD else "",
                "",
                "",
                obj.salesman.get_full_name() or obj.salesman.phone_number,
            ])
        else:
            ws.append([
                obj.date.strftime("%d.%m.%Y"),
                "To'lov",
                obj.note or "",
                "",
                "",
                "",
                float(obj.amount) if obj.currency == UZS else "",
                float(obj.amount) if obj.currency == USD else "",
                obj.created_by.get_full_name() or obj.created_by.phone_number,
            ])
    return _excel_response(wb, f"yetkazib-beruvchi-{supplier.pk}.xlsx")


@accountant_required
def expenditure_export(request):
    qs = (
        Expenditure.objects.filter(is_active=True)
        .select_related("financial_account", "registered_by")
        .order_by("-date", "-created_at")
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Xarajatlar"
    ws.append(["Sana", "Izoh", "Hisob", "Summa", "Valyuta", "Qo'shdi"])
    for exp in qs:
        ws.append([
            exp.date.strftime("%d.%m.%Y"),
            exp.description,
            exp.financial_account.name,
            float(exp.amount),
            exp.currency,
            exp.registered_by.get_full_name() or exp.registered_by.phone_number,
        ])
    return _excel_response(wb, "xarajatlar.xlsx")
